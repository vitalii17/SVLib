import sys
import os

from enum      import Enum, unique
from operator  import itemgetter
from itertools import groupby

from openpyxl        import load_workbook, Workbook
from openpyxl.styles import Font

project_name = "ProjName"
project_ver  = "v1.0"

def main(args):
    if len(args) > 1:
        pass
    else:
        script_path = args[0]
        script_dir  = os.path.dirname(script_path)
        bom_name    = project_name + " " + project_ver + " - BOM"

    raw_bom_reader = XlReader("_raw_bom.xlsx")
    raw_bom = raw_bom_reader.raw_data[1:]
    parser = RawBomParser()
    grouped = parser.separate(raw_bom, 6)
    raw_bom_dict = {}
    raw_bom_dict = parser.separate(grouped, 7)

    bom = Bom(raw_bom_dict)

    xlwriter = XlWriter(file="BOM.xlsx")
    xlwriter.write_bom(bom.structured_bom)


@unique
class RowType(Enum):
    BOMITEM  = 1
    HEADER   = 2
    CATEGORY = 3
    SPACER   = 4


class BomRow(object):
    def __init__(self, cels=(), row_type=RowType.BOMITEM, font_bold=False, 
                                                          font_italic=False, 
                                                          font_underline="none",
                                                          font_size_pt=11):
        super(BomRow, self).__init__()
        self.cels           = cels
        self.row_type       = row_type
        self.font_bold      = font_bold
        self.font_italic    = font_italic
        self.font_underline = font_underline
        self.font_size_pt   = font_size_pt
        

class Bom(object):
    def __init__(self, grouped_bom):
        super(Bom, self).__init__()

        self.header      = BomRow(cels=("Component Name", "Value", "Package", "Count", "Designator", "Notes"),
                                  row_type=RowType.HEADER, font_bold=True, font_size_pt=12)
        self.cat_smd_top = BomRow(cels=("SMD, top side", ""), 
                                  row_type=RowType.CATEGORY, 
                                  font_bold=True, font_italic=True, font_underline="single", font_size_pt=12)
        self.cat_smd_bot = BomRow(cels=("SMD, bottom side", ""), 
                                  row_type=RowType.CATEGORY, 
                                  font_bold=True, font_italic=True, font_underline="single", font_size_pt=12)
        self.cat_pth_top = BomRow(cels=("PTH, top side", ""), 
                                  row_type=RowType.CATEGORY, 
                                  font_bold=True, font_italic=True, font_underline="single", font_size_pt=12)
        self.cat_pth_bot = BomRow(cels=("PTH, bottom side", ""), 
                                  row_type=RowType.CATEGORY, 
                                  font_bold=True, font_italic=True, font_underline="single", font_size_pt=12)
        self.empty_row   = BomRow(cels=("", ""), 
                                  row_type=RowType.SPACER)     
        self.bom_width   = len(self.header.cels)

        self.smd_top = self._make_bom_row_list(grouped_bom.get("SMD", "").get("Top", ""))
        self.smd_bot = self._make_bom_row_list(grouped_bom.get("SMD", "").get("Bottom", ""))
        self.pth_top = self._make_bom_row_list(grouped_bom.get("PTH", "").get("Top", ""))
        self.pth_bot = self._make_bom_row_list(grouped_bom.get("PTH", "").get("Bottom", ""))

        self.structured_bom = [self.header] + \
                              [self.cat_smd_top] + self.smd_top + [self.empty_row] + \
                              [self.cat_smd_bot] + self.smd_bot + [self.empty_row] + \
                              [self.cat_pth_top] + self.pth_top + [self.empty_row] + \
                              [self.cat_pth_bot] + self.pth_bot

    def _make_bom_row_list(self, source):
        result = []
        for item in source:
            result = result + [BomRow(cels=item, row_type=RowType.BOMITEM)]
        return result


class RawBomParser(object):
    def __init__(self):
        super(RawBomParser, self).__init__()

    def separate(self, content, key):
        result_dict = {}
        if isinstance(content, list):
            sorted      = self._sort(content, key)
            result_dict = self._group(sorted, key)
        elif isinstance(content, dict):
            for item in content:
                result_dict[item] = self._group(content[item], key)
        return result_dict
        
    def _sort(self, content, key):
        result = content.copy()
        result.sort(key=itemgetter(key))
        return result

    def _group(self, content, key):
        grouped = groupby(content, key=itemgetter(key))
        result = {}
        for key, items in grouped:
            result_list = []
            for i in items:
                result_list = result_list + [i, ]
            result[key] = result_list
        return result


class XlWriter(object):
    def __init__(self, file):
        super(XlWriter, self).__init__()
        self.file = file

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet1"

    def write_bom(self, row_list):
        for row_num, row_content in enumerate(row_list, start=1):
            self._write_row(row_content, row_num)
        self.wb.save(filename=self.file)

    def _write_row(self, bom_row, row_num):
        for col, val in enumerate(bom_row.cels, start=1):
            self.ws.cell(row=row_num, column=col).value = val
            self.ws.cell(row=row_num, column=col).font  = Font(bold=bom_row.font_bold,
                                                                italic=bom_row.font_italic,
                                                                underline=bom_row.font_underline,
                                                                size=bom_row.font_size_pt)        
        

class XlReader(object):
    def __init__(self, file):
        super(XlReader, self).__init__()
        self.file = file
        self.raw_data = []

        wb = load_workbook(filename=self.file)
        ws = wb.worksheets[0]

        for row in ws.iter_rows(min_row=1):
            raw_row = []
            for item in row:
                raw_row = raw_row + [item.value, ]
            self.raw_data = self.raw_data + [raw_row, ]



if __name__ == '__main__':
    main(sys.argv)
